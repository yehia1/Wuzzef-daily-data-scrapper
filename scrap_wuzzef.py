from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
import time
import pandas as pd
from datetime import date
from openpyxl import load_workbook
import os
import warnings
import argparse
warnings.simplefilter("ignore")

def set_min_max_experience(driver,min:int = 0,max:int = 5):
    
    experience_section = driver.find_element(By.XPATH, "//h3[contains(., 'Years of experience')]")
    experience_section.click()

    time.sleep(1)

    # Wait for the "Min" input field to appear
    wait = WebDriverWait(driver, 10)
    
    # Locate the 'Min' field using the class 'css-1ph4zhu-placeholder'
    min_input = wait.until(EC.presence_of_element_located((By.XPATH, "//div[contains(@class,'css-1ph4zhu-placeholder') and text()='Min']/following::input[1]")))
    min_input.click()
    min_input.clear()
    min_input.send_keys("0")
    min_input.send_keys(Keys.ENTER)


    # Wait for the results page to load
    time.sleep(3)


    # Step 3: Wait for and select the 'Max' input field
    max_input = wait.until(EC.presence_of_element_located((By.XPATH, "//div[contains(@class,'css-1ph4zhu-placeholder') and text()='Max']/following::input[1]")))
    max_input.click()
    max_input.clear()
    max_input.send_keys("5")
    max_input.send_keys(Keys.ENTER)

    print(f"Successfully selected the Years of experience Filter.")
    
def select_date_posted_option(driver,option_text):
    """
    Selects one of the options in the 'Date Posted' filter section on a webpage.
    
    Parameters:
        option_text (str): The text of the option you want to select. 
                           Valid options: 'All', 'Past 24 hours', 'Past week', 'Past month'
    """
    try:

        # Step 1: Open the "Date Posted" section (if not already opened)
        date_posted_section = driver.find_element(By.XPATH, "//h3[contains(., 'Date Posted')]")

        date_posted_section.click()

        driver.execute_script("arguments[0].scrollIntoView(true);", date_posted_section)

        
        # Wait for the "Min" input field to appear
        wait = WebDriverWait(driver, 10)

        # Step 3: Find the radio button using the provided text (e.g., 'Past week')
        option_xpath = f"//span[text()='{option_text}']/preceding::input[1]"
        desired_option = wait.until(EC.presence_of_element_located((By.XPATH, option_xpath)))

        # Step 4: Use JavaScript to click the radio option
        driver.execute_script("arguments[0].click();", desired_option)

        print(f"Successfully selected the '{option_text}' option.")
    
    except Exception as e:
        print(f"An error occurred: {str(e)}")

def extract_job_details(driver, job_data):
    """
    Extracts job details from a list of job cards on a Wuzzuf-like job site and navigates through the job details page.

    Parameters:
    - driver: Selenium WebDriver instance.
    - job_data: List to store the extracted job information. Each job's details will be appended as a dictionary.

    Returns:
    - None: The function adds the scraped job data to the job_data list.
    """

    # Step 1: Find all job cards on the current page
    job_cards = driver.find_elements(By.CLASS_NAME, 'css-1gatmva')
    # Step 2: Iterate through each job card and extract the job details
    for job_card in job_cards:
        try:
            # Find the job title link and its URL
            job_title_link = job_card.find_element(By.CSS_SELECTOR, "a.css-o171kl")
            job_url = job_title_link.get_attribute("href")

            # Open the job URL in a new tab
            driver.execute_script(f"window.open('{job_url}', '_blank');")
                
            # Switch to the new tab where job details are displayed
            driver.switch_to.window(driver.window_handles[-1])
            # Optional: Add a delay to allow the page to load the job details
            time.sleep(2)

            # Step 3: Extract job details (adjust the CSS selectors as per the page's structure)
            job_title = driver.find_element(By.CSS_SELECTOR, 'h1.css-f9uh36').text.strip()
            # Extract company name and location using JavaScript
            company_name = driver.execute_script("return document.querySelector('strong.css-9geu3q').textContent").split('\xa0')[-2][:-2]
            location = driver.execute_script("return document.querySelector('strong.css-9geu3q').textContent").split('\xa0')[-1]
            # Extract experience, salary, skills, description, and job requirements
            experience = driver.find_element(By.CSS_SELECTOR, 'span.css-4xky9y').text.strip()
            salary = driver.find_element(By.XPATH, "//div[contains(@class, 'css-rcl8e5')]//span[contains(text(), 'Salary')]/following-sibling::span").text
            skills = driver.find_element(By.CSS_SELECTOR, 'div.css-s2o0yh').text.split()[3:]
            job_description = driver.find_element(By.CSS_SELECTOR, 'div.css-1uobp1k').text.split('\n')[1:]
            try :
                job_requirements = driver.find_element(By.CSS_SELECTOR, 'div.css-1t5f0fr').text.split('\n')[1:]
            except Exception as e:
                job_requirements = None
            # Append the extracted details to the job_data list
            job_data.append({
                "Job Title": job_title,
                "Company": company_name,
                "Location": location,
                "Job Link": job_url,
                "Salary": salary,
                "Experience": experience,
                "Skills": skills,
                "Job Requirements": job_requirements,
                "Description": job_description
            })

            # Close the job details tab
            driver.close()
            
            # Step 4: Switch back to the original tab to continue extracting the next job
            driver.switch_to.window(driver.window_handles[0])
        except Exception as e:
            # If any error occurs, print it for debugging
            print(f"Failed to extract details for job: {e}")

def navigate_pages(driver, job_data):
    """
    Navigates through the pages of job listings, extracts job details from each page,
    and logs the current page number.

    Args:
        driver: Selenium WebDriver instance.
        job_data: A list to append extracted job details.

    Returns:
        None
    """
    page_index = 1  # Start with the first page

    while True:
        print(f"Processing page {page_index}...")  # Log the current page
        
        # Step 1: Extract job details from the current page
        extract_job_details(driver, job_data)

        try:
            next_button = driver.find_element(By.XPATH, f"//button[contains(@class,'css-zye1os') and .//a[contains(@href,'start={page_index}')]]")
            next_button.click()

            # Optional: Add a delay to allow the next page to load
            time.sleep(3)

            # Step 4: Ensure the page has actually advanced (log the current URL to verify)
            current_url = driver.current_url
            print(f"Successfully moved to page {page_index + 1}. URL: {current_url}")

            # Increment the page index
            page_index += 1

        except Exception as e:
            # If the "Next" button is not found, we are on the last page
            print("No more pages to navigate. Scraping complete.")
            break  # Exit the loop when there are no more pages

def append_sheet_to_excel(df, file_name, sheet_name):
    """
    Append a DataFrame to a new or existing sheet in an Excel file.
    If the file doesn't exist, it creates a new file.
    If the sheet doesn't exist, it creates a new sheet without touching existing sheets.

    Args:
    df (pd.DataFrame): DataFrame to write to Excel.
    file_name (str): Path to the Excel file.
    sheet_name (str): Sheet name to write data to.
    """
    # Check if the file exists
    if os.path.exists(file_name):
        # Load the existing workbook
        book = load_workbook(file_name)
        # Use 'openpyxl' engine to preserve existing sheets
        with pd.ExcelWriter(file_name, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            writer.book = book
            # If the sheet doesn't exist, create it
            if sheet_name not in writer.book.sheetnames:
                df.to_excel(writer, sheet_name=sheet_name, index=False)

                worksheet = writer.sheets[sheet_name]
                
                for row_num, job_link in enumerate(df["Job Link"], start=worksheet.max_row - len(df)):
                    worksheet.cell(row=row_num + 1, column=4).hyperlink = job_link
            else:
                # Get the last row in the existing data
                startrow = writer.sheets[sheet_name].max_row
                # Append the new data below the existing data
                df.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=startrow)

                worksheet = writer.sheets[sheet_name]
                # Iterate over the rows in the DataFrame for adding hyperlinks
                for row_num, job_link in enumerate(df["Job Link"], start=worksheet.max_row - len(df)):  # Adjust start row
                    worksheet.cell(row=row_num + 1, column=4).hyperlink = job_link  # Column 4 is for Job Link

        print(f"Data appended to '{sheet_name}' in '{file_name}'.")
    else:
        # If the file doesn't exist, create a new one
        with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
        print(f"File '{file_name}' created with new sheet '{sheet_name}'.")


# %%
def main(params):

    position = params.position
    posting_time = params.time

    driver = webdriver.Chrome()

    driver.get("https://wuzzuf.net")
    search_input = driver.find_element(By.CLASS_NAME,'css-ukkbbr.e1n2h7jb1')

    # Clear the search input if necessary and enter the job title (e.g., 'Data Analyst')
    search_input.clear()
    search_input.send_keys(position)

    # Simulate pressing the ENTER key to search
    search_input.send_keys(Keys.RETURN)

    # Wait for the results page to load
    time.sleep(5)
    
    set_min_max_experience(driver,min = 0,max =  5)

    # Wait for the results page to load
    time.sleep(3)
    
    # Example usage: Select 'Past week'
    select_date_posted_option(driver,posting_time)
    
    time.sleep(5)

    job_data = []
    navigate_pages(driver,job_data)

    df = pd.DataFrame(job_data)
    append_sheet_to_excel(df, 'Wuzzef_jobs_details.xlsx', sheet_name=f'{date.today()}')


if __name__ == '__main__':
    #Parsing arguments 
    parser = argparse.ArgumentParser(description='Scraping data from wuzzef')

    parser.add_argument('--position', help='position to search for')
    parser.add_argument('--time', help='the time the job was posted')

    args = parser.parse_args()
    main(args)


# py scrap_wuzzef.py --position='Business Intelligence' --time='Past 24 hours'